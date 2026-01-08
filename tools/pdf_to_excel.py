import fitz
import os
from openpyxl import Workbook

def pdf_to_excel(input_path, output_path):
    try:
        doc = fitz.open(input_path)
        wb = Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        for page_num, page in enumerate(doc):
            ws = wb.create_sheet(title=f"Page {page_num + 1}")
            
            # Try to find tables first (requires newer PyMuPDF)
            try:
                tabs = page.find_tables()
                if tabs.tables:
                    current_row = 1
                    for tab in tabs:
                        for row in tab.extract():
                            for col_idx, cell in enumerate(row, 1):
                                # Clean cell content
                                content = str(cell).strip() if cell else ""
                                ws.cell(row=current_row, column=col_idx, value=content)
                            current_row += 1
                        current_row += 2 # Spacer between tables
                else:
                    # Fallback to text blocks if no tables found
                    blocks = page.get_text("blocks")
                    blocks.sort(key=lambda b: (b[1], b[0])) # Sort by Y then X
                    
                    for i, block in enumerate(blocks, 1):
                        text = block[4].strip()
                        # Simple logic: put each block in A column, or try to respect position
                        # For robustness, we'll just list text blocks in column A for now
                        # improving this requires complex layout analysis
                        ws.cell(row=i, column=1, value=text)
            except Exception as e:
                # Fallback purely to blocks if table extraction fails/isn't supported
                blocks = page.get_text("blocks")
                blocks.sort(key=lambda b: (b[1], b[0]))
                for i, block in enumerate(blocks, 1):
                    ws.cell(row=i, column=1, value=block[4].strip())

        wb.save(output_path)
        doc.close()
        
        return {'success': True, 'output_path': output_path, 'filename': os.path.basename(output_path)}
    except Exception as e:
        return {'success': False, 'error': str(e)}
