from openpyxl import load_workbook
import os
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

def excel_to_pdf(input_path, output_path):
    try:
        wb = load_workbook(input_path, data_only=True)
        
        pdf = SimpleDocTemplate(output_path, pagesize=landscape(letter),
                               rightMargin=36, leftMargin=36,
                               topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        story = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            story.append(Paragraph(f"Sheet: {sheet_name}", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            data = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                if any(row_data):
                    data.append(row_data)
            
            if data:
                max_cols = max(len(row) for row in data)
                for i, row in enumerate(data):
                    while len(row) < max_cols:
                        data[i] = list(row) + ['']
                        row = data[i]
                
                col_width = min(700 / max_cols, 150)
                col_widths = [col_width] * max_cols
                
                table = Table(data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(table)
            
            story.append(Spacer(1, 24))
        
        pdf.build(story)
        
        return {'success': True, 'output_path': output_path, 'filename': os.path.basename(output_path)}
    except Exception as e:
        return {'success': False, 'error': str(e)}
